##from FUNCTIONS import *
from openpyxl import load_workbook

##pathTradeExport = r"C:\Users\KDK\Desktop\ALLINONE.xlsx"
##df = pd.read_excel(pathTradeExport)

##ws = wb.worksheets[2]
##"TICKER","B",
pathTradeExport = r"C:\Users\KDK\Desktop\14.11-17.11\14.11-17.11.xlsx"

def prepareTitles(ws, pathTradeExport):
    titles = ["A", "C", "D", "E", "F"]
    values = ["DATE", "TYPE", "PRICE", "AMOUNT", "TOTAL"]
    column = 1
    position = 0
    wb = load_workbook(pathTradeExport)
    sheet_number = len(wb.worksheets)
    print(sheet_number)
    for _ in values:
                ##sheet_number = sheet_number - 1
                ##ws = wb.worksheets[sheet_number]
                ws.cell(row=1, column=column).value = values[position]
                ws.column_dimensions[titles[position]].width = 22


                ##print(position)
                ##print(column)
                print(sheet_number)

                ##sheet_number = sheet_number - 1
                column = column + 1
                position = position + 1
                ##wb.save(pathTradeExport)


    wb.save(pathTradeExport)

def prepareTitlesAllSheets(pathTradeExport):
    wb = load_workbook(pathTradeExport)
    sheet_number = len(wb.worksheets)
    for ws in wb.worksheets:
        sheet_number=sheet_number-1
        ws = wb.worksheets[sheet_number]
        prepareTitles(ws, pathTradeExport)
        wb.save(pathTradeExport)

prepareTitlesAllSheets(pathTradeExport)