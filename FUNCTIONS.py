from openpyxl import load_workbook


pathTradeExport = r"C:\Users\KDK\Desktop\FOR testing binc3.xlsx"

def getDate(x, pathTradeExport):         ##class v ktery se bude dedit ws = wb wrksheets ?
    wb = load_workbook(pathTradeExport)
    ws = wb.worksheets[0]
    date = ws.cell(row=x, column=1).value
    ##print(date)
    return (date)


def getTicker(x, pathTradeExport):                          ##getTickerName.. x=row parametr x, column=Market , BTCBUSD
        wb = load_workbook(pathTradeExport)
        ws = wb.worksheets[0]
        ticker = ws.cell(row=x, column=2).value
        ##print(ticker)
        return(ticker)


def getTypeoftrade(x, pathTradeExport):                  #sell or buy
    wb = load_workbook(pathTradeExport)
    ws = wb.worksheets[0]
    Typeoftrade = ws.cell(row=x, column=3).value
    print(Typeoftrade)
    return (Typeoftrade)

def getPrice(x, pathTradeExport):                  #price
    wb = load_workbook(pathTradeExport)
    ws = wb.worksheets[0]
    price = ws.cell(row=x, column=4).value
    print(price)
    return (price)


def getAmount(x,pathTradeExport):                  #amount
    wb = load_workbook(pathTradeExport)
    ws = wb.worksheets[0]
    amount = ws.cell(row=x, column=5).value
    print(amount)
    return (amount)

def getTotalpaid(x, pathTradeExport):                  #totalPaid
    wb = load_workbook(pathTradeExport)
    ws = wb.worksheets[0]
    total = ws.cell(row=x, column=6).value
    print(total)
    return (total)
x=4


def getAllInfo(x, pathTradeExport):
    getDate(x, pathTradeExport)
    getTicker(x, pathTradeExport)
    getTypeoftrade(x, pathTradeExport)
    getPrice(x, pathTradeExport)
    getAmount(x, pathTradeExport)
    getTotalpaid(x, pathTradeExport)

getAllInfo(x, pathTradeExport)

