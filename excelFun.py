import pandas as pd
import time
import openpyxl
from selenium import webdriver
from openpyxl import load_workbook, cell

pathTradeExport = r"C:\Users\KDK\Desktop\FOR testing binc.xlsx"
ticker = "dot"
df = pd.read_excel(pathTradeExport)
wb = load_workbook(pathTradeExport)
ws = wb.worksheets[0]


##starting everytime on row 2 based on bnc export
numeroOfRows = ws.max_row
x= 2

while numeroOfRows > 1:
  kekw = ws.cell(row=x, column=2).value
  print(kekw)
  x = x+1
  print(x)
  numeroOfRows = numeroOfRows-1
  print(numeroOfRows)

wb.sheetnames()

wb.create_sheet(ticker)         ##udelej sheet ticker, switchni donej a placni value
ws = wb.worksheets[ticker]


ws.cell(row=1, column=11).value = "REDIRECT STATUS"
wb.save(pathTradeExport)

##ws.cell(row=2, column="Market").value
##print(d)

##print(numeroOfRows)



##ws.cell(row=1, column=11).value = "REDIRECT STATUS"

