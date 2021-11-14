from openpyxl import load_workbook
import time
pathTradeExport = r"C:\Users\KDK\Desktop\ALLINONE.xlsx"
pathTradeExport = r"C:\Users\KDK\Desktop\EXCELY BINANCE ALL\1.9 - 14.11 2021 ALL TRADES.xlsx"
##pathTradeExport = r"C:\Users\KDK\Desktop\pairSellBuy.xlsx"
##this gets total sold and total bought which is cool but not what i wanted to do in the first place
x=2 ##STARTING ROW

wb = load_workbook(pathTradeExport)
ws = wb.worksheets[1]

amount = ws.cell(row=x, column=4).value
numeroOfRows = ws.max_row

Typeoftrade = ws.cell(row=x, column=2).value

sell = "SELL"
sellsCount = 0
buysCount = 0
sellsAmount = 0
buysAmount = 0
sellsPrices = []
buysPrices = []
sellsAmounts = []
buysAmounts = []
while numeroOfRows > 1:
    Typeoftrade = ws.cell(row=x, column=2).value
    total = ws.cell(row=x, column=5).value

    totalInt = int(float(total))
    price = ws.cell(row=x, column=3).value
    priceInt = float((price))

    amount = ws.cell(row=x, column=4).value
    amountInt = float(amount)

    print(total)
    ##print(Typeoftrade)

    if Typeoftrade == "SELL":
        ##print("je to sell")
        sellsCount = sellsCount+1
        ##print(sellsCount)
        sellsPrices.append(priceInt)
        sellsAmounts.append(amountInt)

        print(sellsPrices )
        print(sellsAmounts)
        numeroOfRows = numeroOfRows - 1
        sellsAmount = sellsAmount+totalInt
        print(sum(sellsAmounts))
        x=x+1

    if Typeoftrade == "BUY":
        buysAmounts.append(amountInt)

        buysPrices.append(priceInt)
        buysCount = buysCount+1



        print(buysPrices)
        print(buysAmounts)
        numeroOfRows = numeroOfRows - 1
        buysAmount = buysAmount + totalInt
        print(sum(buysAmounts))
        x=x+1


    else:
        pass
        print("pass")


buysAmountsSUM = sum(buysAmounts)
print(buysAmountsSUM)

sellsAmountsSUM = sum(sellsAmounts)
print(sellsAmountsSUM)

ws.cell(row=2, column=8).value = sellsCount
ws.cell(row=2, column=9).value = buysCount
ws.cell(row=2, column=10).value = sellsAmount
ws.cell(row=2, column=11).value = buysAmount

wb.save(pathTradeExport)
