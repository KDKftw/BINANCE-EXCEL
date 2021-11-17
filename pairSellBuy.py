from openpyxl import load_workbook
from delete_empty_rows import deleteInAllSheets
pathTradeExport = r"C:\Users\KDK\Desktop\pairSellBuy.xlsx"
pathTradeExport = r"C:\Users\KDK\Desktop\14.11-17.11\14.11-17.11.xlsx"
x=2 ##STARTING ROW

wb = load_workbook(pathTradeExport)
ws = wb.worksheets[0]  ##variable when to iritate



numeroOfRows = ws.max_row

print(numeroOfRows)

def getAllRowValues(sellPuvodniRow, sellPresunuteRow):
    dateTrade = ws.cell(row=sellPuvodniRow, column=1).value
    Typeoftrade = ws.cell(row=sellPuvodniRow, column=2).value
    Priceoftrade = ws.cell(row=sellPuvodniRow, column=3).value
    amount = ws.cell(row=sellPuvodniRow, column=4).value
    totalTrade = ws.cell(row=sellPuvodniRow, column=5).value

    ##dostanu vsechyn values co potrebuju

    ws.cell(row=sellPresunuteRow, column=1).value = dateTrade
    ws.cell(row=sellPresunuteRow, column=2).value = Typeoftrade
    ws.cell(row=sellPresunuteRow, column=3).value = Priceoftrade
    ws.cell(row=sellPresunuteRow, column=4).value = amount
    ws.cell(row=sellPresunuteRow, column=5).value = totalTrade
##23-31

sellSortingStartRow = numeroOfRows + 2
while numeroOfRows > 0:
    Typeoftrade = ws.cell(row=numeroOfRows, column=2).value
    if Typeoftrade == "SELL":
      getAllRowValues(numeroOfRows, sellSortingStartRow )
      ws.delete_rows(numeroOfRows, 1)
      sellSortingStartRow = sellSortingStartRow+1
      numeroOfRows = numeroOfRows - 1

    else:
        print("pass")
        numeroOfRows = numeroOfRows - 1


wb.save(pathTradeExport)




