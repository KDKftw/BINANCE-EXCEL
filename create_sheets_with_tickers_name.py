from SetupFunctions import *
from openpyxl import load_workbook

pathTradeExport3 = r"C:\Users\KDK\Desktop\ALLINONE.xlsx"
pathTradeExport2 = r"C:\Users\KDK\Desktop\12.6 - 1.8 2021 ALL TRADES2.xlsx"
pathTradeExport = r"C:\Users\KDK\Desktop\BINANCE_EXCEL_START-14.11.2021\excel_all2.xlsx"
wb = load_workbook(pathTradeExport)
ws = wb.worksheets[0]

##starting everytime on row 2 based on bnc export
def createSheetsTickers(pathTradeExport):
    numeroOfRows = ws.max_row
    x=2
    while numeroOfRows > 1:
      tickerName = ws.cell(row=x, column=2).value
      if tickerName in (wb.sheetnames):
          pass
      else:
          wb.create_sheet(tickerName)
      x = x+1
      numeroOfRows = numeroOfRows-1
      wb.save(pathTradeExport)

createSheetsTickers(pathTradeExport)

