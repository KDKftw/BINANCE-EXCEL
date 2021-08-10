import pandas as pd
import time
import openpyxl
from selenium import webdriver
from openpyxl import load_workbook, cell
from FUNCTIONS import *

pathTradeExport = r"C:\Users\KDK\Desktop\FOR testing binc.xlsx"
df = pd.read_excel(pathTradeExport)
wb = load_workbook(pathTradeExport)
ws = wb.worksheets[1]
ws.cell(row=x, column=6).value= getTotalpaid(x)
wb.save(pathTradeExport)
