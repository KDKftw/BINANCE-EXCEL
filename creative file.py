from SetupFunctions import *
from openpyxl import load_workbook

pathTradeExport = r"C:\Users\KDK\Desktop\FOR testing binc3.xlsx"
pathTradeExport2 = r"C:\Users\KDK\Desktop\12.6 - 1.8 2021 ALL TRADES2.xlsx"
wb = load_workbook(pathTradeExport)
ws = wb.worksheets[0]

##starting everytime on row 2 based on bnc export
def createSheetsTickers():
    numeroOfRows = ws.max_row
    x=2
    while numeroOfRows > 1:
      tickerName = ws.cell(row=x, column=2).value
      if tickerName in (wb.sheetnames):
          pass
      else:
          wb.create_sheet(tickerName)
      x = x+1
      numeroOfRows = numeroOfRows-1
      wb.save(pathTradeExport)

createSheetsTickers()


def copyValuesToSheets():
    ws = wb.worksheets[0]
    numeroOfRows = ws.max_row
    print(numeroOfRows)
    x = 2
    while numeroOfRows > 1:
        sheetPosition = wb.worksheets.index(wb[getTicker(x)])
        ws = wb.worksheets[sheetPosition]
        ws.cell(row=x, column=1).value = getDate(x)

        x = x + 1
        numeroOfRows = numeroOfRows - 1
        wb.save(pathTradeExport)
    wb.save(pathTradeExport)

##copyValuesToSheets()


