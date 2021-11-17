from openpyxl import load_workbook

def deleteExtraRows(ws, pathTradeExport):
    wb = load_workbook(pathTradeExport)
    index_row = []
    # loop each row in column A
    for i in range(1, ws.max_row):
        # define emptiness of cell
        if ws.cell(i, 1).value is None:
            # collect indexes of rows
            index_row.append(i)

    # loop each index value
    for row_del in range(len(index_row)):
        ws.delete_rows(idx=index_row[row_del], amount=1)
        # exclude offset of rows through each iteration
        index_row = list(map(lambda k: k - 1, index_row))
        wb.save(pathTradeExport)


 ##x is one cuz we want to start on the second sheet, since the first one is where all the data is
def deleteInAllSheets(pathTradeExport):
    x = 1
    wb = load_workbook(pathTradeExport)
    sheet_number = len(wb.worksheets)
    print(sheet_number)
    while sheet_number > 1:
        ws = wb.worksheets[x]
        deleteExtraRows(ws, pathTradeExport)

        x=x+1
        sheet_number = sheet_number-1
        print(x)
        wb.save(pathTradeExport)

pathTradeExport = r"C:\Users\KDK\Desktop\14.11-17.11\14.11-17.11.xlsx"
deleteInAllSheets(pathTradeExport)