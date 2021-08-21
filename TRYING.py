import openpyxl
from openpyxl import load_workbook
pathTradeExport2 = r"C:\Users\KDK\Desktop\FOR testing binc.xlsx"
pathTradeExport = r"C:\Users\KDK\Desktop\FOR testing binc3.xlsx"
from FUNCTIONS import getPrice, getDate, getTypeoftrade, getTicker, getTotalpaid, getAmount




def copyValuesToSheets():
        wb = load_workbook(pathTradeExport)
        ws = wb.worksheets[0]
        numeroOfRows = ws.max_row
        print(numeroOfRows)
        x = 2
        while numeroOfRows > 1:
            sheetPosition = wb.worksheets.index(wb[getTicker(x, pathTradeExport)])
            ws = wb.worksheets[sheetPosition]
            ws.cell(row=x, column=1).value = getDate(x, pathTradeExport)

            sheetPosition = wb.worksheets.index(wb[getTicker(x, pathTradeExport)])
            ws = wb.worksheets[sheetPosition]
            ws.cell(row=x, column=2).value = getTypeoftrade(x, pathTradeExport)

            sheetPosition = wb.worksheets.index(wb[getTicker(x, pathTradeExport)])
            ws = wb.worksheets[sheetPosition]
            ws.cell(row=x, column=3).value = getPrice(x, pathTradeExport)

            sheetPosition = wb.worksheets.index(wb[getTicker(x, pathTradeExport)])
            ws = wb.worksheets[sheetPosition]
            ws.cell(row=x, column=4).value = getAmount(x, pathTradeExport)

            sheetPosition = wb.worksheets.index(wb[getTicker(x, pathTradeExport)])
            ws = wb.worksheets[sheetPosition]
            ws.cell(row=x, column=5).value = getTotalpaid(x, pathTradeExport)

            x = x + 1
            numeroOfRows = numeroOfRows - 1
            wb.save(pathTradeExport)
        wb.save(pathTradeExport)


copyValuesToSheets()